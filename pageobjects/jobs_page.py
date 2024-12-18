
import datetime
import os
from selenium.webdriver.common.by import By
from utilities import XLUtils
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from pageobjects.locators import LocatorsHomePage
from pageobjects.locators import LocatorsMenu
from pageobjects.locators import LocatorsCompany
from pageobjects.locators import LocatorsJobs
from allure_commons.types import AttachmentType
import allure
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import xlsxwriter
from selenium.webdriver.common.alert import Alert
from selenium.common.exceptions import TimeoutException 
from selenium.common.exceptions import NoAlertPresentException
from selenium import webdriver
import pyautogui
import shutil
from openpyxl import Workbook
import time
import xlsxwriter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

import openpyxl 

class JobsPage():
    
    path=r"test_data\BBS6.xlsx"

    def __init__(self,driver):
        self.driver=driver

    def create_excel(self):
        # workbook=xlsxwriter.Workbook(r"C:\Nagalakshmi\PythonPractice\BBS_automation\test_data\BBS.xlsx")
        # worksheet = workbook.add_worksheet("Jobs")
        # worksheet.write(0, 0, "TEXT FOR ASSERTION")
        # worksheet.write(0, 1, "RESULTS")
        # worksheet.write(2, 0, "JOBS")
        # worksheet.write(3, 0,"All Job Openings")

        # workbook.close()

        wb = Workbook()
        file_path = "test_data"
        create_duplicate = 'BixBytes website' + str(datetime.datetime.now().strftime('%d-%m-%Y %H-%M-%S')) + '.xlsx'
        replica = os.path.join(file_path, create_duplicate)
        wb.save(replica)
        source = "test_data\BBS6.xlsx"
        dest = replica
        shutil.copy(source, dest)

#================================================================JOBS PAGE================================================================
    def click_hamburger(self):
        self.driver.find_element(*LocatorsHomePage.btn_Hamburger_hmpg_x).click()

    def menu_jobs_link(self):
        menu_jobs_link=XLUtils.readData(self.path,"Jobs",3,2)
        print(menu_jobs_link)
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(LocatorsMenu.menu_jobs_link))
        print((self.driver.find_element(*LocatorsMenu().menu_jobs_link)).text)
        if (self.driver.find_element(*LocatorsMenu().menu_jobs_link)).text == menu_jobs_link:
            XLUtils.writeData(self.path,"Jobs",3,3,"pass")
            XLUtils.fillGreenColor(self.path,"Jobs",3,3)
            assert True
        else:
            XLUtils.writeData(self.path,"Jobs",3,3,"failed")
            XLUtils.fillRedColor(self.path,"Jobs",3,3)
            self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
            allure.attach(self.driver.get_screenshot_as_png(), name="dashboard_tc_fail", attachment_type=AttachmentType.PNG)
            assert False
        time.sleep(2)
        Jobs=self.driver.find_element(*LocatorsMenu.menu_jobs_link)
        self.driver.execute_script('arguments[0].click()', Jobs)
        time.sleep(2)

    def scroll_down_Page(self):
        # self.driver.refresh()
        # logo= WebDriverWait(self.driver,20).until(EC.visibility_of_element_located(LocatorsJobs.logo_jobs_page_x))
        # logo.click()
        self.driver.find_element(By.XPATH, "//h2[normalize-space()='IT Project Lead (PHP)']").click()
        for i in range(0,12):
            if i<12:
                 ActionChains(self.driver).send_keys(Keys.SPACE).perform()
                 time.sleep(1)

    def scroll_up_Page(self):
        for i in range(0,12):
            if i<12:
                 ActionChains(self.driver).send_keys(Keys.PAGE_UP).perform()
                 time.sleep(1)
        hamburger=self.driver.find_element(*LocatorsJobs.hamburger_jobs_page_x)
        self.driver.execute_script('arguments[0].click()', hamburger)
        time.sleep(2)
        self.driver.refresh()

    def scroll_stop_between(self):
        txt= WebDriverWait(self.driver,20).until(EC.visibility_of_element_located(LocatorsJobs.logo_jobs_page_x))
        txt.click()
        for i in range(0,3):
            if i<5:
                 ActionChains(self.driver).send_keys(Keys.PAGE_DOWN).perform()
                 time.sleep(1)

    def jobs_12read_more(self):
        twelve_readmore=WebDriverWait(self.driver,20).until(EC.presence_of_all_elements_located(LocatorsJobs.jobs_12read_more))
        for each_link in range(len(twelve_readmore)):
            #Added this line to avoid stale which re-assigned the element again.
            twelve_readmore=self.driver.find_elements(By.XPATH, "//div[@class='job_article_main mt-60 position-relative']//a[@class='read_more fontSize-2_4 font-weight-400']")
            # six_readmore[each_link].click()
            self.driver.execute_script("arguments[0].click();", twelve_readmore[each_link])
            time.sleep(2)
            img_clk=self.driver.find_element(By.XPATH,"//div[@class='job_details_bg_img']")
            img_clk.click()
            ActionChains(self.driver).send_keys(Keys.SPACE).perform()
            ActionChains(self.driver).send_keys(Keys.SPACE).perform()
            time.sleep(2)
            self.driver.back()
            time.sleep(2)

    def txt_all_job_openings(self):
        txt_all_job_openings=XLUtils.readData(self.path,"Jobs",4,2)
        print(txt_all_job_openings)
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(LocatorsJobs.txt_all_job_openings))
        print((self.driver.find_element(*LocatorsJobs().txt_all_job_openings)).text)
        if (self.driver.find_element(*LocatorsJobs().txt_all_job_openings)).text == txt_all_job_openings:
            XLUtils.writeData(self.path,"Jobs",4,3,"pass")
            XLUtils.fillGreenColor(self.path,"Jobs",4,3)
            assert True
        else:
            XLUtils.writeData(self.path,"Jobs",4,3,"failed")
            XLUtils.fillRedColor(self.path,"Jobs",4,3)
            self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
            allure.attach(self.driver.get_screenshot_as_png(), name="dashboard_tc_fail", attachment_type=AttachmentType.PNG)
            assert False
            
    def all_openings_12job_cards(self):
        # logo= WebDriverWait(self.driver,20).until(EC.presence_of_element_located(LocatorsJobs.job_page_1line))
        # logo.click()
        # for i in range(0,6):
        #     if i<6:
        #          ActionChains(self.driver).send_keys(Keys.PAGE_UP).perform()
        #          time.sleep(2)
        twelve_readmore=WebDriverWait(self.driver,20).until(EC.presence_of_all_elements_located(LocatorsJobs.all_openings_job_cards))
        for each_link in range(len(twelve_readmore)):
            #Added this line to avoid stale which re-assigned the element again.
            twelve_readmore=self.driver.find_elements(By.XPATH, "//a[@class='read_more font-weight-4000 fontSize-2_0 cursor-pointer text-uppercase']")
            # six_readmore[each_link].click()
            self.driver.execute_script("arguments[0].click();", twelve_readmore[each_link])
            time.sleep(2)
            img_clk=self.driver.find_element(By.XPATH,"//div[@class='job_details_bg_img']")
            img_clk.click()
            ActionChains(self.driver).send_keys(Keys.SPACE).perform()
            ActionChains(self.driver).send_keys(Keys.SPACE).perform()
            time.sleep(2)
            self.driver.back()
            time.sleep(2)

 #================================================================JOB DETAIL PAGE===============================================================#
    def jobs_read_more_12th(self):
        ActionChains(self.driver).send_keys(Keys.ARROW_UP).perform()
        jobs_read_more_12th=WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(LocatorsJobs.jobs_read_more_12th))
        self.driver.execute_script("arguments[0].click();", jobs_read_more_12th)

    def asrt_job_detpag_hdline(self):
        asrt_job_detpag_hdline=XLUtils.readData(self.path,"Jobs",17,2)
        print(asrt_job_detpag_hdline)
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(LocatorsJobs.asrt_job_detpag_hdline))
        print((self.driver.find_element(*LocatorsJobs().asrt_job_detpag_hdline)).text)
        if (self.driver.find_element(*LocatorsJobs.asrt_job_detpag_hdline)).text == asrt_job_detpag_hdline:
            XLUtils.writeData(self.path,"Jobs",17,3,"pass")
            XLUtils.fillGreenColor(self.path,"Jobs",17,3)
            assert True
        else:
            XLUtils.writeData(self.path,"Jobs",17,3,"failed")
            XLUtils.fillRedColor(self.path,"Jobs",17,3)
            self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
            allure.attach(self.driver.get_screenshot_as_png(), name="dashboard_tc_fail", attachment_type=AttachmentType.PNG)
            assert False
        time.sleep(5)  
    
    def asrt_job_detpag_txt1(self):
        list_data=[]
        list_eachEle=[]
       
        four_elements=WebDriverWait(self.driver, 10).until(EC.presence_of_all_elements_located(LocatorsJobs.asrt_job_detpag_txt1))
        for each_element in four_elements:
            list_eachEle.append(each_element.text)
        print(list_eachEle, "all elements names:")
        
        for r in range(18,20):
            data=XLUtils.readData(self.path,"Jobs",r,2)
            list_data.append(data)
        print(list_data, "Excel data:")
        
        for i in range(0,len(list_data)) :
            for r in range(18,20):
                if list_data[i]==list_eachEle[i]:
                    XLUtils.writeData(self.path,"Jobs",r,3,"pass")
                    XLUtils.fillGreenColor(self.path,"Jobs",r,3)
                    assert True
                else:
                    XLUtils.writeData(self.path,"Jobs",r,3,"failed")
                    XLUtils.fillRedColor(self.path,"Jobs",r,3)
                    self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
                    allure.attach(self.driver.get_screenshot_as_png(), name="tc_fail", attachment_type=AttachmentType.PNG)
                    assert False

    def asrt_job_detpag_txt2(self):
        list_data=[]
        list_eachEle=[]
       
        four_elements=WebDriverWait(self.driver, 10).until(EC.presence_of_all_elements_located(LocatorsJobs.asrt_job_detpag_txt2))
        for each_element in four_elements:
            list_eachEle.append(each_element.text)
        print(list_eachEle, "all elements names:")
        
        for r in range(20,22):
            data=XLUtils.readData(self.path,"Jobs",r,2)
            list_data.append(data)
        print(list_data, "Excel data:")
        
        for i in range(0,len(list_data)) :
            for r in range(20,22):
                if list_data[i]==list_eachEle[i]:
                    XLUtils.writeData(self.path,"Jobs",r,3,"pass")
                    XLUtils.fillGreenColor(self.path,"Jobs",r,3)
                    assert True
                else:
                    XLUtils.writeData(self.path,"Jobs",r,3,"failed")
                    XLUtils.fillRedColor(self.path,"Jobs",r,3)
                    self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
                    allure.attach(self.driver.get_screenshot_as_png(), name="tc_fail", attachment_type=AttachmentType.PNG)
                    assert False

    def asrt_job_detpag_txt3(self):
        list_data=[]
        list_eachEle=[]
       
        four_elements=WebDriverWait(self.driver, 10).until(EC.presence_of_all_elements_located(LocatorsJobs.asrt_job_detpag_txt3))
        for each_element in four_elements:
            list_eachEle.append(each_element.text)
        print(list_eachEle, "all elements names:")
        
        for r in range(22,26):
            data=XLUtils.readData(self.path,"Jobs",r,2)
            list_data.append(data)
        print(list_data, "Excel data:")
        
        for i in range(0,len(list_data)):
            for r in range(22,26):
                if list_data[i]==list_eachEle[i]:
                    XLUtils.writeData(self.path,"Jobs",r,3,"pass")
                    XLUtils.fillGreenColor(self.path,"Jobs",r,3)
                    assert True
                else:
                    XLUtils.writeData(self.path,"Jobs",r,3,"failed")
                    XLUtils.fillRedColor(self.path,"Jobs",r,3)
                    self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
                    allure.attach(self.driver.get_screenshot_as_png(), name="tc_fail", attachment_type=AttachmentType.PNG)
                    assert False

    def asrt_job_detpag_txt4(self):
        list_data=[]
        list_eachEle=[]
       
        four_elements=WebDriverWait(self.driver, 10).until(EC.presence_of_all_elements_located(LocatorsJobs.asrt_job_detpag_txt4))
        for each_element in four_elements:
            list_eachEle.append(each_element.text)
        print(list_eachEle, "all elements names:")
        
        for r in range(26,28):
            data=XLUtils.readData(self.path,"Jobs",r,2)
            list_data.append(data)
        print(list_data, "Excel data:")
        
        for i in range(0,len(list_data)):
            for r in range(26,28):
                if list_data[i]==list_eachEle[i]:
                    XLUtils.writeData(self.path,"Jobs",r,3,"pass")
                    XLUtils.fillGreenColor(self.path,"Jobs",r,3)
                    assert True
                else:
                    XLUtils.writeData(self.path,"Jobs",r,3,"failed")
                    XLUtils.fillRedColor(self.path,"Jobs",r,3)
                    self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
                    allure.attach(self.driver.get_screenshot_as_png(), name="tc_fail", attachment_type=AttachmentType.PNG)
                    assert False

    def jop_application_heading(self):
        jop_application_heading=XLUtils.readData(self.path,"Jobs",28,2)
        print(jop_application_heading)
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(LocatorsJobs.jop_application_heading))
        print((self.driver.find_element(*LocatorsJobs().jop_application_heading)).text)
        if (self.driver.find_element(*LocatorsJobs.jop_application_heading)).text == jop_application_heading:
            XLUtils.writeData(self.path,"Jobs",28,3,"pass")
            XLUtils.fillGreenColor(self.path,"Jobs",28,3)
            assert True
        else:
            XLUtils.writeData(self.path,"Jobs",28,3,"failed")
            XLUtils.fillRedColor(self.path,"Jobs",28,3)
            self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
            allure.attach(self.driver.get_screenshot_as_png(), name="dashboard_tc_fail", attachment_type=AttachmentType.PNG)
            assert False

    def job_aplcn_first_name(self):
        job_aplcn_first_name=WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(LocatorsJobs.job_aplcn_first_name))
        job_aplcn_first_name.send_keys("abc")

    def job_aplcn_last_name(self):
        job_aplcn_last_name=WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(LocatorsJobs.job_aplcn_last_name))
        job_aplcn_last_name.send_keys("xyz")
    
    def job_aplcn_email(self):
        job_aplcn_email=WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(LocatorsJobs.job_aplcn_email))
        job_aplcn_email.send_keys("xyz@gmail.com")
        # job_aplcn_email.send_keys("xyz")

    def job_aplcn_phone(self):
        job_aplcn_phone=WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(LocatorsJobs.job_aplcn_phone))
        job_aplcn_phone.send_keys("+918217098759")
        # job_aplcn_phone.send_keys("123")

    def upload_resume(self):
        upload_resume=WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(LocatorsJobs.upload_resume))
        # self.driver.execute_script("arguments[0].click();",upload_resume)
        input=self.driver.find_element(By.XPATH, "//input[@type='file']")
        input.send_keys(r"C:\Users\NagalakshmiS\Desktop\BBS0139-NAGALAKSHMI.pdf")
        # pyautogui.write(r'C:\Users\NagalakshmiS\Desktop\my files\ChallengesC:\Users\NagalakshmiS\Desktop\New Text Document.txt')
        time.sleep(2) 
        # pyautogui.press('Esc')
        # parent_window = print(self.driver.current_window_handle, "#value of parent window")
        # child_windows=self.driver.window_handles
        # print(child_windows, "#value of child windows")
        # for each_window in child_windows:
        #     self.driver.switch_to.window(each_window)
        #     print(self.driver.title, "#title of all windows")
        #     time.sleep(1)
        #     if self.driver.title!="SharePoint Software Engineer | Jobs - Bix Bytes Solutions":
        #        self.driver.close()
        #        time.sleep(3)

    def job_aplcn_portfolio(self):
        job_aplcn_portfolio=WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(LocatorsJobs.job_aplcn_portfolio))
        job_aplcn_portfolio.send_keys("aaaa")

    def click_ch_box(self):
        click_ch_box=WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(LocatorsJobs.upload_resume))
        self.driver.execute_script("arguments[0].click();",click_ch_box)

    def job_aplcn_submit_btn(self):
        job_aplcn_submit_btn=WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(LocatorsJobs.job_aplcn_submit_btn))
        self.driver.execute_script("arguments[0].click();",job_aplcn_submit_btn)
        time.sleep(2)
        self.driver.back()

    def warning_msg1_aplcn_form(self):
        list_data=[]
        list_eachEle=[]
       
        two_elements=WebDriverWait(self.driver, 10).until(EC.presence_of_all_elements_located(LocatorsJobs.warning_msg1_aplcn_form))
        for each_element in two_elements:
            list_eachEle.append(each_element.text)
        print(list_eachEle, "all elements names:")
        
        for r in range(31,33):
            data=XLUtils.readData(self.path,"Jobs",r,2)
            list_data.append(data)
        print(list_data, "Excel data:")
        
        for i in range(0,len(list_data)):
            for r in range(31,33):
                if list_data[i]==list_eachEle[i]:
                    XLUtils.writeData(self.path,"Jobs",r,3,"pass")
                    XLUtils.fillGreenColor(self.path,"Jobs",r,3)
                    assert True
                else:
                    XLUtils.writeData(self.path,"Jobs",r,3,"failed")
                    XLUtils.fillRedColor(self.path,"Jobs",r,3)
                    self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
                    allure.attach(self.driver.get_screenshot_as_png(), name="tc_fail", attachment_type=AttachmentType.PNG)
                    assert False

    def warning_msg_last_name(self):
        warning_msg_last_name=XLUtils.readData(self.path,"Jobs",33,2)
        print(warning_msg_last_name)
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(LocatorsJobs.warning_msg_last_name))
        print((self.driver.find_element(*LocatorsJobs().warning_msg_last_name)).text)
        if (self.driver.find_element(*LocatorsJobs.warning_msg_last_name)).text == warning_msg_last_name:
            XLUtils.writeData(self.path,"Jobs",33,3,"pass")
            XLUtils.fillGreenColor(self.path,"Jobs",33,3)
            assert True
        else:
            XLUtils.writeData(self.path,"Jobs",33,3,"failed")
            XLUtils.fillRedColor(self.path,"Jobs",33,3)
            self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
            allure.attach(self.driver.get_screenshot_as_png(), name="dashboard_tc_fail", attachment_type=AttachmentType.PNG)
            assert False

    def warning_msg_phnum(self):
        warning_msg_phnum=XLUtils.readData(self.path,"Jobs",34,2)
        print(warning_msg_phnum)
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(LocatorsJobs.warning_msg_phnum))
        print((self.driver.find_element(*LocatorsJobs().warning_msg_phnum)).text)
        if (self.driver.find_element(*LocatorsJobs.warning_msg_phnum)).text == warning_msg_phnum:
            XLUtils.writeData(self.path,"Jobs",34,3,"pass")
            XLUtils.fillGreenColor(self.path,"Jobs",34,3)
            assert True
        else:
            XLUtils.writeData(self.path,"Jobs",34,3,"failed")
            XLUtils.fillRedColor(self.path,"Jobs",34,3)
            self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
            allure.attach(self.driver.get_screenshot_as_png(), name="dashboard_tc_fail", attachment_type=AttachmentType.PNG)
            assert False

    def warning_msg_uploadfile(self):
        warning_msg_uploadfile=XLUtils.readData(self.path,"Jobs",35,2)
        print(warning_msg_uploadfile)
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(LocatorsJobs.warning_msg_uploadfile))
        print((self.driver.find_element(*LocatorsJobs().warning_msg_uploadfile)).text)
        if (self.driver.find_element(*LocatorsJobs.warning_msg_uploadfile)).text == warning_msg_uploadfile:
            XLUtils.writeData(self.path,"Jobs",35,3,"pass")
            XLUtils.fillGreenColor(self.path,"Jobs",35,3)
            assert True
        else:
            XLUtils.writeData(self.path,"Jobs",35,3,"failed")
            XLUtils.fillRedColor(self.path,"Jobs",35,3)
            self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
            allure.attach(self.driver.get_screenshot_as_png(), name="dashboard_tc_fail", attachment_type=AttachmentType.PNG)
            assert False
        time.sleep(2)

    def warning_msg_captcha(self):
        warning_msg_captcha=XLUtils.readData(self.path,"Jobs",36,2)
        print(warning_msg_captcha)
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(LocatorsJobs.warning_msg_captcha))
        print((self.driver.find_element(*LocatorsJobs().warning_msg_captcha)).text)
        if (self.driver.find_element(*LocatorsJobs.warning_msg_captcha)).text == warning_msg_captcha:
            XLUtils.writeData(self.path,"Jobs",36,3,"pass")
            XLUtils.fillGreenColor(self.path,"Jobs",36,3)
            assert True
        else:
            XLUtils.writeData(self.path,"Jobs",36,3,"failed")
            XLUtils.fillRedColor(self.path,"Jobs",36,3)
            self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
            allure.attach(self.driver.get_screenshot_as_png(), name="dashboard_tc_fail", attachment_type=AttachmentType.PNG)
            assert False

    def error_msg_invalid_mailid(self):
        error_msg_invalid_mailid=XLUtils.readData(self.path,"Jobs",37,2)
        print(error_msg_invalid_mailid)
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(LocatorsJobs.error_msg_invalid_mailid))
        print((self.driver.find_element(*LocatorsJobs().error_msg_invalid_mailid)).text)
        if (self.driver.find_element(*LocatorsJobs.error_msg_invalid_mailid)).text == error_msg_invalid_mailid:
            XLUtils.writeData(self.path,"Jobs",37,3,"pass")
            XLUtils.fillGreenColor(self.path,"Jobs",37,3)
            assert True
        else:
            XLUtils.writeData(self.path,"Jobs",37,3,"failed")
            XLUtils.fillRedColor(self.path,"Jobs",37,3)
            self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
            allure.attach(self.driver.get_screenshot_as_png(), name="dashboard_tc_fail", attachment_type=AttachmentType.PNG)
            assert False

    def error_msg_invalid_phn_num(self):
        error_msg_invalid_phn_num=XLUtils.readData(self.path,"Jobs",38,2)
        print(error_msg_invalid_phn_num)
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(LocatorsJobs.error_msg_invalid_phn_num))
        print((self.driver.find_element(*LocatorsJobs().error_msg_invalid_phn_num)).text)
        if (self.driver.find_element(*LocatorsJobs.error_msg_invalid_phn_num)).text == error_msg_invalid_phn_num:
            XLUtils.writeData(self.path,"Jobs",38,3,"pass")
            XLUtils.fillGreenColor(self.path,"Jobs",38,3)
            assert True
        else:
            XLUtils.writeData(self.path,"Jobs",38,3,"failed")
            XLUtils.fillRedColor(self.path,"Jobs",38,3)
            self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
            allure.attach(self.driver.get_screenshot_as_png(), name="dashboard_tc_fail", attachment_type=AttachmentType.PNG)
            assert False

    def jop_aplcn_footer_txt(self):
        jop_aplcn_footer_txt=XLUtils.readData(self.path,"Jobs",39,2)
        print(jop_aplcn_footer_txt)
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(LocatorsJobs.jop_aplcn_footer_txt))
        print((self.driver.find_element(*LocatorsJobs().jop_aplcn_footer_txt)).text)
        if (self.driver.find_element(*LocatorsJobs.jop_aplcn_footer_txt)).text == jop_aplcn_footer_txt:
            XLUtils.writeData(self.path,"Jobs",39,3,"pass")
            XLUtils.fillGreenColor(self.path,"Jobs",39,3)
            # assert True
        else:
            XLUtils.writeData(self.path,"Jobs",39,3,"failed")
            XLUtils.fillRedColor(self.path,"Jobs",39,3)
            self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
            allure.attach(self.driver.get_screenshot_as_png(), name="dashboard_tc_fail", attachment_type=AttachmentType.PNG)
            # assert False

    def merge_cells(self):
        
        wb = Workbook()
        file_path = "test_data"
        create_duplicate = 'BixBytes website' + str(datetime.datetime.now().strftime('%d-%m-%Y %H-%M-%S')) + '.xlsx'
        replica = os.path.join(file_path, create_duplicate)
        wb.save(replica)
        source = "test_data\BBS6.xlsx"
        dest = replica
        shutil.copy(source, dest)

        print("TTTT :", source , " YYYY :", dest)
 
        print("Copied and created")
 
        # file=dest
        # # file = r"C:\Nagalakshmi\PythonPractice\BBS_automation\test_data\BixBytes website.xlsx"
        # print("FFFFFFFFFFFFF", file)
       
        workbook = openpyxl.load_workbook(dest)
        worksheet = workbook['Jobs']
        indicators_excel = []

        for row_num in range(7, 15):
              cell = worksheet.cell(row_num, column=3)
              indicators_excel.append(cell.value)
        print("Appended values : ", indicators_excel)

        worksheet.merge_cells('D7:D14')
        worksheet['D7'].alignment = Alignment(horizontal='center', vertical='center')
        if "failed" in indicators_excel:
            worksheet['D7'].alignment = Alignment(horizontal='center', vertical='center')
            worksheet['D7'] = "Failed"
            top_left_cell = worksheet['D7']
            top_left_cell.fill  = PatternFill(start_color='ff0000', end_color='ff0000',
                                        fill_type = "solid")
        else:
            worksheet['D7'].alignment = Alignment(horizontal='center', vertical='center')
            worksheet['D7'] = "Passed"
            top_left_cell = worksheet['D7']
            top_left_cell.fill  = PatternFill(start_color='60b212', end_color='60b212',
                                        fill_type = "solid")
        workbook.save(dest)

#================================================================FOOTER================================================================#

    def footer_headline(self):
        footer_headline=XLUtils.readData(self.path,"Jobs",7,2)
        print(footer_headline)
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(LocatorsHomePage.footer_headline))
        print((self.driver.find_element(*LocatorsHomePage().footer_headline)).text)
        if (self.driver.find_element(*LocatorsHomePage.footer_headline)).text == footer_headline:
            XLUtils.writeData(self.path,"Jobs",7,3,"pass")
            XLUtils.fillGreenColor(self.path,"Jobs",7,3)
            assert True
        else:
            XLUtils.writeData(self.path,"Jobs",7,3,"failed")
            XLUtils.fillRedColor(self.path,"Jobs",7,3)
            self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
            allure.attach(self.driver.get_screenshot_as_png(), name="dashboard_tc_fail", attachment_type=AttachmentType.PNG)
            assert False

    def footer_assert_switzerland(self):
        footer_assert_switzerland=XLUtils.readData(self.path,"Jobs",8,2)
        print(footer_assert_switzerland)
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(LocatorsHomePage.footer_assert_switzerland))
        print((self.driver.find_element(*LocatorsHomePage().footer_assert_switzerland)).text)
        if (self.driver.find_element(*LocatorsHomePage.footer_assert_switzerland)).text == footer_assert_switzerland:
            XLUtils.writeData(self.path,"Jobs",8,3,"pass")
            XLUtils.fillGreenColor(self.path,"Jobs",8,3)
            assert True
        else:
            XLUtils.writeData(self.path,"Jobs",8,3,"failed")
            XLUtils.fillRedColor(self.path,"Jobs",8,3)
            self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
            allure.attach(self.driver.get_screenshot_as_png(), name="dashboard_tc_fail", attachment_type=AttachmentType.PNG)
            assert False

    def footer_assert_br_Ind_mang(self):
        footer_assert_br_Ind_mang=XLUtils.readData(self.path,"Jobs",9,2)
        print(footer_assert_br_Ind_mang)
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(LocatorsHomePage.footer_assert_br_Ind_mang))
        print((self.driver.find_element(*LocatorsHomePage().footer_assert_br_Ind_mang)).text)
        if (self.driver.find_element(*LocatorsHomePage.footer_assert_br_Ind_mang)).text == footer_assert_br_Ind_mang:
            XLUtils.writeData(self.path,"Jobs",9,3,"pass")
            XLUtils.fillGreenColor(self.path,"Jobs",9,3)
            assert True
        else:
            XLUtils.writeData(self.path,"Jobs",9,3,"failed")
            XLUtils.fillRedColor(self.path,"Jobs",9,3)
            self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
            allure.attach(self.driver.get_screenshot_as_png(), name="dashboard_tc_fail", attachment_type=AttachmentType.PNG)
            assert False

    def footer_assert_br_canada(self):
        footer_assert_br_canada=XLUtils.readData(self.path,"Jobs",10,2)
        print(footer_assert_br_canada)
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(LocatorsHomePage.footer_assert_br_canada))
        print((self.driver.find_element(*LocatorsHomePage().footer_assert_br_canada)).text)
        if (self.driver.find_element(*LocatorsHomePage().footer_assert_br_canada)).text == footer_assert_br_canada:
            XLUtils.writeData(self.path,"Jobs",10,3,"pass")
            XLUtils.fillGreenColor(self.path,"Jobs",10,3)
            assert True
        else:
            XLUtils.writeData(self.path,"Jobs",10,3,"failed")
            XLUtils.fillRedColor(self.path,"Jobs",10,3)
            self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
            allure.attach(self.driver.get_screenshot_as_png(), name="dashboard_tc_fail", attachment_type=AttachmentType.PNG)
            assert False

    def footer_assert_br_Ind_bang(self):
        footer_assert_br_Ind_bang=XLUtils.readData(self.path,"Jobs",11,2)
        print(footer_assert_br_Ind_bang)
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(LocatorsHomePage.footer_assert_br_Ind_bang))
        print((self.driver.find_element(*LocatorsHomePage().footer_assert_br_Ind_bang)).text)
        if (self.driver.find_element(*LocatorsHomePage.footer_assert_br_Ind_bang)).text == footer_assert_br_Ind_bang:
            XLUtils.writeData(self.path,"Jobs",11,3,"pass")
            XLUtils.fillGreenColor(self.path,"Jobs",11,3)
            assert True
        else:
            XLUtils.writeData(self.path,"Jobs",11,3,"failed")
            XLUtils.fillRedColor(self.path,"Jobs",11,3)
            self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
            allure.attach(self.driver.get_screenshot_as_png(), name="dashboard_tc_fail", attachment_type=AttachmentType.PNG)
            assert False

    def footer_assert_company_name(self):
        footer_assert_company_name=XLUtils.readData(self.path,"Jobs",12,2)
        print(footer_assert_company_name)
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(LocatorsHomePage.footer_assert_company_name))
        print((self.driver.find_element(*LocatorsHomePage().footer_assert_company_name)).text)
        if (self.driver.find_element(*LocatorsHomePage.footer_assert_company_name)).text == footer_assert_company_name:
            XLUtils.writeData(self.path,"Jobs",12,3,"pass")
            XLUtils.fillGreenColor(self.path,"Jobs",12,3)
            assert True
        else:
            XLUtils.writeData(self.path,"Jobs",12,3,"failed")
            XLUtils.fillRedColor(self.path,"Jobs",12,3)
            self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
            allure.attach(self.driver.get_screenshot_as_png(), name="dashboard_tc_fail", attachment_type=AttachmentType.PNG)
            assert False
            time.sleep(2)

    def footer_mailaddress_company_page(self):
        footer_mailaddress_company_page=XLUtils.readData(self.path,"Jobs",13,2)
        print(footer_mailaddress_company_page)
        WebDriverWait(self.driver, 30).until(EC.visibility_of_element_located(LocatorsCompany.footer_mailaddress_company_page))
        print((self.driver.find_element(*LocatorsCompany().footer_mailaddress_company_page)).text)
        if (self.driver.find_element(*LocatorsCompany.footer_mailaddress_company_page)).text == footer_mailaddress_company_page:
            XLUtils.writeData(self.path,"Jobs",13,3,"pass")
            XLUtils.fillGreenColor(self.path,"Jobs",13,3)
            assert True
        else:
            XLUtils.writeData(self.path,"Jobs",13,3,"failed")
            XLUtils.fillRedColor(self.path,"Jobs",13,3)
            self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
            allure.attach(self.driver.get_screenshot_as_png(), name="dashboard_tc_fail", attachment_type=AttachmentType.PNG)
            assert False
            time.sleep(2)

    def footer_end_text(self):
        footer_end_text=XLUtils.readData(self.path,"Jobs",14,2)
        print(footer_end_text)
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(LocatorsHomePage.footer_end_text))
        print((self.driver.find_element(*LocatorsHomePage().footer_end_text)).text)
        if (self.driver.find_element(*LocatorsHomePage.footer_end_text)).text == footer_end_text:
            XLUtils.writeData(self.path,"Jobs",14,3,"pass")
            XLUtils.fillGreenColor(self.path,"Jobs",14,3)
            # assert True
        else:
            XLUtils.writeData(self.path,"Jobs",14,3,"failed")
            XLUtils.fillRedColor(self.path,"Jobs",14,3)
            self.driver.save_screenshot(r"C:\\Nagalakshmi\\PythonPractice\\BBS_automation\\screenshots\\failed_tc.png")
            allure.attach(self.driver.get_screenshot_as_png(), name="dashboard_tc_fail", attachment_type=AttachmentType.PNG)
            # assert False

    def click_footer_insta_link(self):
        footer_mail_link=self.driver.find_element(By.XPATH, "//div[@class='row mb-lg-130 mb-md-70 mb-40 lg_flagged_information']//div[1]//div[1]//div[1]//p[2]//a[1]")
        for i in range(0,12):
            if i<12:
                 ActionChains(self.driver).send_keys(Keys.SPACE).perform()
        footer_insta_link= WebDriverWait(self.driver,20).until(EC.visibility_of_element_located(LocatorsHomePage.footer_insta_link))
        self.driver.execute_script("arguments[0].click();",footer_insta_link)
        time.sleep(1)
    
    def click_footer_fb_link(self):
        footer_fb_link= WebDriverWait(self.driver,20).until(EC.visibility_of_element_located(LocatorsHomePage.footer_fb_link))
        self.driver.execute_script("arguments[0].click();",footer_fb_link)
        time.sleep(1)

    def click_footer_twitter_link(self):
        footer_twitter_link= WebDriverWait(self.driver,20).until(EC.visibility_of_element_located(LocatorsHomePage.footer_twitter_link))
        self.driver.execute_script("arguments[0].click();",footer_twitter_link)
        time.sleep(1)

    def click_footer_linkedin_link(self):
        footer_linkedin_link= WebDriverWait(self.driver,20).until(EC.visibility_of_element_located(LocatorsHomePage.footer_linkedin_link))
        self.driver.execute_script("arguments[0].click();",footer_linkedin_link)
        time.sleep(1)

    def close_child_windows(self):
        print(self.driver.current_window_handle, "#value of parent window")
        child_windows=self.driver.window_handles
        print(child_windows, "#value of child windows")
        for each_window in child_windows:
            self.driver.switch_to.window(each_window)
            print(self.driver.title, "#title of all windows")
            time.sleep(1)
            if self.driver.title!="Automation QA Software Engineer | Jobs - Bix Bytes Solutions":
               self.driver.close()
        time.sleep(3)
        self.driver.switch_to.window(self.driver.window_handles[0])

    def click_footer_mail_address_link(self):
        # footer_mail_link=self.driver.find_element(By.XPATH, "//div[@class='row mb-lg-130 mb-md-70 mb-40 lg_flagged_information']//div[1]//div[1]//div[1]//p[2]//a[1]")
        # for i in range(0,12):
        #     if i<12:
        #          ActionChains(self.driver).send_keys(Keys.SPACE).perform()
        footer_mail_link= WebDriverWait(self.driver,20).until(EC.visibility_of_element_located(LocatorsJobs.footer_mail_link))
        self.driver.execute_script("arguments[0].click();",footer_mail_link)
        time.sleep(6)

    def close_mail_window(self):
        pyautogui.hotkey('alt','F4')
        time.sleep(2)
        pyautogui.hotkey('PAGE_DOWN','ENTER')

   
    def footer_contact_num_link_cp(self):
        footer_contact_link= WebDriverWait(self.driver,20).until(EC.visibility_of_element_located(LocatorsJobs.footer_contact_link))
        self.driver.execute_script("arguments[0].click();",footer_contact_link)
        time.sleep(3)

    def clear_pop_up(self):
        pyautogui.press('Esc')
        time.sleep(3) 
       
   

       
       






    
   
